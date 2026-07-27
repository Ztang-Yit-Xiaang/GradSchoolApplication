"""
Advisor Finder — Excel Builder
===============================
Usage:
    python build_advisor_excel.py --state ADVISOR_STATE.md --output advisors.xlsx

Or import and call build_workbook() directly with structured data dicts.

This script generates a multi-sheet workbook matching the Advisor Finder spec:
  Sheet 1: Weighted Ranking (formula totals, color by priority)
  Sheet 2: Detailed Profile comparison
  Sheet 3: Outreach Priority & Angles
  Sheet 4: Email Templates
  Sheet 5: Methodology, Weights, Recency Log, Sources
  Sheet 6: Entry Papers per top advisor
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("openpyxl not found. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Color palette ──────────────────────────────────────────────────────────────
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")   # top 5 recruitable
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")   # next tier
GRAY_FILL   = PatternFill("solid", fgColor="D9D9D9")   # needs confirm
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")    # filtered / ❌
HEADER_FILL = PatternFill("solid", fgColor="2F5496")   # dark blue header
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BOLD        = Font(bold=True)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="top")


def _header_row(ws, headers: list[str], row: int = 1):
    """Write a styled header row."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Sheet 1: Weighted Ranking ──────────────────────────────────────────────────
def build_sheet1(wb, advisors: list[dict], interests: list[dict]):
    """
    advisors: list of dicts with keys:
        rank, name, school, dept, title, homepage_updated, recency_source,
        scores (dict: area -> 0-10), weighted_total, recruiting, priority, oneliner
    interests: list of dicts with keys: area, weight (normalized, sums to 1.0)
    """
    ws = wb.active
    ws.title = "1_加权评分排序"

    areas   = [i["area"]   for i in interests]
    weights = [i["weight"] for i in interests]

    # Row 1 = weights reference row (hidden or labeled)
    ws.cell(row=1, column=1, value="权重行（勿删）").font = Font(italic=True, color="808080")
    for col, w in enumerate(weights, start=8):  # scores start at col 8
        ws.cell(row=1, column=col, value=w)

    # Row 2 = column headers
    headers = (
        ["排名", "导师", "学校", "院系/职称", "主页更新", "数据源"]
        + [f"{a}\n(w={w:.2f})" for a, w in zip(areas, weights)]
        + ["加权总分", "招生状态", "优先级", "一句话理由"]
    )
    _header_row(ws, headers, row=2)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    score_start_col = 7  # column index (1-based) where scores begin
    n_areas = len(areas)
    total_col = score_start_col + n_areas        # weighted total column
    recruit_col = total_col + 1
    priority_col = recruit_col + 1
    oneliner_col = priority_col + 1

    for i, adv in enumerate(advisors):
        row = i + 3
        scores = adv.get("scores", {})

        ws.cell(row=row, column=1, value=adv.get("rank", i + 1))
        ws.cell(row=row, column=2, value=adv.get("name", ""))
        ws.cell(row=row, column=3, value=adv.get("school", ""))
        ws.cell(row=row, column=4, value=f"{adv.get('dept','')}\n{adv.get('title','')}").alignment = WRAP
        ws.cell(row=row, column=5, value=adv.get("homepage_updated", ""))
        ws.cell(row=row, column=6, value=adv.get("recency_source", "homepage"))

        # Sub-scores
        score_cells = []
        for j, area in enumerate(areas):
            col = score_start_col + j
            val = scores.get(area, 0)
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = CENTER
            score_cells.append(get_column_letter(col) + str(row))

        # Weighted total formula: =SUMPRODUCT(scores_range, weights_range)
        score_range   = f"{get_column_letter(score_start_col)}{row}:{get_column_letter(score_start_col+n_areas-1)}{row}"
        weights_range = f"{get_column_letter(score_start_col)}1:{get_column_letter(score_start_col+n_areas-1)}1"
        ws.cell(row=row, column=total_col,    value=f"=SUMPRODUCT({score_range},{weights_range})").number_format = "0.00"
        ws.cell(row=row, column=recruit_col,  value=adv.get("recruiting", "❓"))
        ws.cell(row=row, column=priority_col, value=adv.get("priority", ""))
        ws.cell(row=row, column=oneliner_col, value=adv.get("oneliner", "")).alignment = WRAP

        # Row color
        status = adv.get("recruiting", "❓")
        priority = adv.get("priority", "")
        if "❌" in status:
            fill = RED_FILL
        elif priority in ("A", "★★★", "Top") or (i < 5 and "✅" in status):
            fill = GREEN_FILL
        elif "✅" in status:
            fill = YELLOW_FILL
        else:
            fill = GRAY_FILL

        for col in range(1, oneliner_col + 1):
            ws.cell(row=row, column=col).fill = fill

    _set_col_widths(ws, [5, 18, 16, 20, 12, 10] + [10] * n_areas + [10, 12, 8, 40])


# ── Sheet 2: Detailed Profiles ─────────────────────────────────────────────────
def build_sheet2(wb, advisors: list[dict]):
    ws = wb.create_sheet("2_详情对照")
    headers = ["导师", "学校/职称", "招生状态", "研究方向", "近三年代表成果", "与候选人契合点", "主页", "邮箱"]
    _header_row(ws, headers)
    ws.freeze_panes = "A2"

    for i, adv in enumerate(advisors, 2):
        ws.cell(row=i, column=1, value=adv.get("name", ""))
        ws.cell(row=i, column=2, value=f"{adv.get('school','')} / {adv.get('title','')}").alignment = WRAP
        ws.cell(row=i, column=3, value=adv.get("recruiting", "❓"))
        ws.cell(row=i, column=4, value=adv.get("directions", "")).alignment = WRAP
        ws.cell(row=i, column=5, value=adv.get("papers", "")).alignment = WRAP
        ws.cell(row=i, column=6, value=adv.get("fit_note", "")).alignment = WRAP
        ws.cell(row=i, column=7, value=adv.get("homepage", ""))
        ws.cell(row=i, column=8, value=adv.get("email", ""))

    _set_col_widths(ws, [18, 22, 12, 40, 50, 35, 35, 28])


# ── Sheet 3: Outreach Priority ────────────────────────────────────────────────
def build_sheet3(wb, advisors: list[dict]):
    ws = wb.create_sheet("3_套磁优先级")
    headers = ["排名", "导师", "学校", "招生状态", "套磁角度", "建议邮件主题"]
    _header_row(ws, headers)
    ws.freeze_panes = "A2"

    for i, adv in enumerate(advisors, 2):
        ws.cell(row=i, column=1, value=adv.get("rank", i - 1))
        ws.cell(row=i, column=2, value=adv.get("name", ""))
        ws.cell(row=i, column=3, value=adv.get("school", ""))
        ws.cell(row=i, column=4, value=adv.get("recruiting", "❓"))
        ws.cell(row=i, column=5, value=adv.get("outreach_angle", "")).alignment = WRAP
        ws.cell(row=i, column=6, value=adv.get("email_subject", "")).alignment = WRAP

    _set_col_widths(ws, [5, 18, 16, 12, 55, 40])


# ── Sheet 4: Email Templates ──────────────────────────────────────────────────
def build_sheet4(wb, templates: dict):
    ws = wb.create_sheet("4_邮件模板")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    rows = [
        ("中文模板", templates.get("zh", "[粘贴中文套磁模板]")),
        ("English Template", templates.get("en", "[Paste English outreach template]")),
        ("个性化清单", templates.get("personalization", "每封邮件务必个性化：导师姓名、一篇具体近作、一个具体研究想法")),
    ]
    for r, (label, content) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=label).font = BOLD
        c = ws.cell(row=r, column=2, value=content)
        c.alignment = WRAP
        ws.row_dimensions[r].height = max(60, content.count("\n") * 15 + 20)


# ── Sheet 5: Methodology ──────────────────────────────────────────────────────
def build_sheet5(wb, meta: dict):
    ws = wb.create_sheet("5_说明·权重·来源")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    sections = [
        ("生成日期",       meta.get("date", "")),
        ("目标学位",       meta.get("degree", "")),
        ("目标范围",       meta.get("target", "")),
        ("权重设定",       meta.get("weights_note", "")),
        ("评分口径",       "9-10=核心主攻; 6-8=活跃次方向; 4-5=相邻可迁移; 2-3=偶有涉及; 0-1=无"),
        ("时效核查日志",   meta.get("recency_log", "见各导师数据源列")),
        ("数据来源",       meta.get("sources", "")),
        ("免责声明",       "招生状态、职称、论文随时变动。发信前务必以导师主页与研究生院官方信息为准。本表不构成录取承诺。打分含主观成分，仅供排序参考。"),
    ]
    for r, (label, content) in enumerate(sections, 1):
        ws.cell(row=r, column=1, value=label).font = BOLD
        c = ws.cell(row=r, column=2, value=str(content))
        c.alignment = WRAP
        ws.row_dimensions[r].height = max(20, str(content).count("\n") * 15 + 20)


# ── Sheet 6: Entry Papers ─────────────────────────────────────────────────────
def build_sheet6(wb, entry_papers: list[dict]):
    """
    entry_papers: list of dicts with keys:
        advisor, title, venue_year, summary, why_fit
    """
    ws = wb.create_sheet("6_入手论文")
    headers = ["导师", "论文标题", "发表/年份", "内容简要", "为何适合候选人"]
    _header_row(ws, headers)
    ws.freeze_panes = "A2"

    for i, paper in enumerate(entry_papers, 2):
        ws.cell(row=i, column=1, value=paper.get("advisor", ""))
        ws.cell(row=i, column=2, value=paper.get("title", "")).alignment = WRAP
        ws.cell(row=i, column=3, value=paper.get("venue_year", ""))
        ws.cell(row=i, column=4, value=paper.get("summary", "")).alignment = WRAP
        ws.cell(row=i, column=5, value=paper.get("why_fit", "")).alignment = WRAP

    _set_col_widths(ws, [18, 45, 15, 45, 40])


# ── Main builder ──────────────────────────────────────────────────────────────
def build_workbook(
    advisors:     list[dict],
    interests:    list[dict],
    entry_papers: list[dict],
    templates:    dict,
    meta:         dict,
    output_path:  str,
):
    """
    Build the full 6-sheet workbook and save to output_path.

    Parameters
    ----------
    advisors : list of advisor dicts (see Sheet 1 docstring for keys)
    interests : list of {area, weight} dicts (normalized weights)
    entry_papers : list of {advisor, title, venue_year, summary, why_fit}
    templates : {zh: str, en: str, personalization: str}
    meta : {date, degree, target, weights_note, recency_log, sources}
    output_path : path for .xlsx output
    """
    wb = openpyxl.Workbook()
    build_sheet1(wb, advisors, interests)
    build_sheet2(wb, advisors)
    build_sheet3(wb, advisors)
    build_sheet4(wb, templates)
    build_sheet5(wb, meta)
    build_sheet6(wb, entry_papers)

    path = Path(output_path)
    wb.save(path)

    # Quick load-verify
    try:
        openpyxl.load_workbook(path)
        print(f"✅ Excel saved and verified: {path}")
    except Exception as e:
        print(f"❌ Verification failed: {e}")
        raise

    return str(path)


# ── CLI entry point ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Advisor Finder Excel workbook")
    parser.add_argument("--output", default="advisors.xlsx", help="Output .xlsx path")
    parser.add_argument("--demo",   action="store_true",     help="Generate a demo workbook with fake data")
    args = parser.parse_args()

    if args.demo:
        # ── Demo data ──
        interests = [
            {"area": "LLM Agent", "weight": 0.40},
            {"area": "Reasoning", "weight": 0.30},
            {"area": "Multimodal", "weight": 0.20},
            {"area": "Medical AI", "weight": 0.10},
        ]
        advisors = [
            {
                "rank": 1, "name": "Prof. Alice Zhang", "school": "HKUST(GZ)",
                "dept": "Information Hub", "title": "Assistant Professor",
                "homepage_updated": "2025-03", "recency_source": "Scholar",
                "scores": {"LLM Agent": 9, "Reasoning": 8, "Multimodal": 6, "Medical AI": 2},
                "recruiting": "✅", "priority": "A",
                "oneliner": "Core LLM agent + reasoning focus; 3 top-tier 2024-25 papers.",
                "directions": "LLM agents, tool use, multi-step reasoning",
                "papers": "[2025] NeurIPS – AgentBench++ (1st author)\n[2024] ICLR – Chain-of-Action (corresponding)",
                "fit_note": "Candidate's RL-for-agents project directly maps to her 2025 AgentBench++ work.",
                "homepage": "https://example.com/alice", "email": "alice@hkust.edu",
                "outreach_angle": "Your 2025 NeurIPS AgentBench++ paper shows agents still struggle with long-horizon planning. My RL-based hierarchical planner (arXiv 2024) addresses exactly this gap — I'd love to explore combining them.",
                "email_subject": "[PhD Inquiry] LLM Agents – Jane Doe",
            },
            {
                "rank": 2, "name": "Prof. Bob Liu", "school": "HKUST(GZ)",
                "dept": "Computational Media Hub", "title": "Associate Professor",
                "homepage_updated": "2023-11", "recency_source": "Scholar (homepage stale)",
                "scores": {"LLM Agent": 6, "Reasoning": 9, "Multimodal": 8, "Medical AI": 1},
                "recruiting": "❓", "priority": "B",
                "oneliner": "Strong reasoning + multimodal; recruiting status unclear for MPhil.",
                "directions": "Visual reasoning, VQA, multimodal LLMs",
                "papers": "[2026] CVPR – VisReason (corresponding)\n[2025] ICLR – MMReason (co-author)",
                "fit_note": "Candidate's vision-language project aligns with his VisReason CVPR 2026 work.",
                "homepage": "https://example.com/bob", "email": "bob@hkust.edu",
                "outreach_angle": "Your CVPR 2026 VisReason paper tackles compositional visual reasoning; my NLI-grounded visual QA work (ACL 2024) could extend your evaluation framework.",
                "email_subject": "[PhD Inquiry] Visual Reasoning – Jane Doe",
            },
        ]
        entry_papers = [
            {
                "advisor": "Prof. Alice Zhang",
                "title": "AgentBench++: Evaluating LLMs as Agents at Scale",
                "venue_year": "NeurIPS 2025",
                "summary": "大规模评测 LLM 作为智能体的基准，涵盖多步工具调用与长程规划。",
                "why_fit": "候选人 RL-based planner 项目与本文评估框架直接对应，读完可找到具体切入点。",
            },
            {
                "advisor": "Prof. Bob Liu",
                "title": "VisReason: Compositional Visual Reasoning in Multimodal LLMs",
                "venue_year": "CVPR 2026",
                "summary": "提出组合式视觉推理框架，显著提升 VQA 精度。",
                "why_fit": "候选人视觉语言项目与此方向高度吻合，是套磁的核心论据。",
            },
        ]
        templates = {
            "zh": "尊敬的 [姓名] 教授，\n\n我是 [候选人]，[学校] [专业] [年级]。\n\n我读了您 [年份] 发表在 [venue] 的「[论文名]」，其中 [具体内容] 给我留下了深刻印象。结合我在 [项目/论文] 中的工作，我认为 [具体研究方向] 有进一步探索的空间。\n\n请问您 [学位季] 是否有招收 [学位] 学生的计划？附上我的简历，期待您的回复。\n\n此致\n[候选人]",
            "en": "Dear Professor [Last Name],\n\nI am [Name], a [Year] student at [School] majoring in [Field].\n\nYour [Year] [Venue] paper \"[Title]\" impressed me, particularly [specific finding]. Building on my work in [project/paper], I see an opportunity to [concrete research direction].\n\nAre you planning to recruit [Degree] students for [Term]? I have attached my CV and would welcome the chance to discuss.\n\nBest regards,\n[Name]",
            "personalization": "每封邮件必须包含：\n1. 导师全名（正确拼写）\n2. 一篇具体近作（标题+venue）\n3. 一个具体研究想法（非泛泛「感兴趣」）\n4. 目标学位与入学季\n5. 简历附件",
        }
        meta = {
            "date": "2026-06-30",
            "degree": "PhD",
            "target": "HKUST(GZ) Information Hub",
            "weights_note": "LLM Agent 0.40 · Reasoning 0.30 · Multimodal 0.20 · Medical AI 0.10 (已归一化)",
            "recency_log": "Prof. Bob Liu: 主页停于 2023-11，改用 Google Scholar 获取近年论文。",
            "sources": "https://facultyprofile.hkust-gz.edu.cn/...\nhttps://scholar.google.com/...",
        }

        build_workbook(
            advisors=advisors,
            interests=interests,
            entry_papers=entry_papers,
            templates=templates,
            meta=meta,
            output_path=args.output,
        )
    else:
        print("No --demo flag. Provide data programmatically via build_workbook() or use --demo to generate sample output.")
